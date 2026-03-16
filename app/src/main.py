from pathlib import Path
from collections import Counter
from datetime import datetime
from html import escape
from io import BytesIO
from urllib.parse import quote

from fastapi import Depends, FastAPI, HTTPException, Request
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from openpyxl import Workbook
from sqlalchemy import desc, func
from sqlalchemy.orm import Session

from app.database import Base, SessionLocal, engine
from app.models import GradeClick, Visitor
from app.schemas import (
    GradeClickRecord,
    GradeClickResponse,
    GradeClickStats,
    VisitResponse,
    VisitorRecord,
)

Base.metadata.create_all(bind=engine)

app = FastAPI(title="Visitor Counter")

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
GRADE_NAMES = ["6 класс", "7 класс", "8 класс", "9 класс", "10 класс", "11 класс"]

def get_client_ip(request: Request) -> str | None:
    cf_ip = request.headers.get("cf-connecting-ip")
    if cf_ip:
        return cf_ip.strip()

    forwarded_for = request.headers.get("x-forwarded-for")
    if forwarded_for:
        # RFC 7239 chain format: client, proxy1, proxy2...
        first_hop = forwarded_for.split(",")[0].strip()
        if first_hop:
            return first_hop

    real_ip = request.headers.get("x-real-ip")
    if real_ip:
        return real_ip.strip()

    return request.client.host if request.client else None


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def validate_grade_name(grade_name: str) -> str:
    if grade_name not in GRADE_NAMES:
        raise HTTPException(status_code=404, detail="Grade not found")
    return grade_name


def render_grade_stats_page(stats: list[GradeClickStats]) -> str:
    rows = "\n".join(
        f"""
        <tr>
          <td>{escape(item.grade_name)}</td>
          <td>{item.clicks}</td>
          <td><a href="/grade-clicks/view/{quote(item.grade_name)}">Открыть</a></td>
        </tr>
        """
        for item in stats
    )

    return f"""<!DOCTYPE html>
<html lang="ru">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Статистика по классам</title>
  <style>
    body {{
      margin: 0;
      font-family: "Trebuchet MS", "Segoe UI", sans-serif;
      background: linear-gradient(135deg, #fff8e8 0%, #eef6ff 100%);
      color: #1f2937;
      padding: 24px;
    }}
    .card {{
      max-width: 900px;
      margin: 0 auto;
      background: rgba(255, 255, 255, 0.94);
      border-radius: 18px;
      padding: 28px;
      box-shadow: 0 18px 40px rgba(34, 54, 84, 0.14);
    }}
    h1 {{
      margin: 0 0 10px;
    }}
    p {{
      color: #5b6472;
      margin: 0 0 20px;
    }}
    .actions {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 16px;
      margin-bottom: 20px;
      flex-wrap: wrap;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      overflow: hidden;
      border-radius: 14px;
      background: #fff;
    }}
    th, td {{
      padding: 14px 16px;
      text-align: left;
      border-bottom: 1px solid #e5e7eb;
    }}
    th {{
      background: #f6efe5;
    }}
    a {{
      color: #9f5314;
      text-decoration: none;
      font-weight: 700;
    }}
  </style>
</head>
<body>
  <div class="card">
    <h1>Статистика по классам</h1>
    <div class="actions">
      <p>Сводка по количеству нажатий на кнопки.</p>
      <a href="/grade-clicks/export">Скачать Excel</a>
    </div>
    <table>
      <thead>
        <tr>
          <th>Класс</th>
          <th>Нажатий</th>
          <th>Детали</th>
        </tr>
      </thead>
      <tbody>
        {rows}
      </tbody>
    </table>
  </div>
</body>
</html>"""


def render_grade_detail_page(grade_name: str, grade_clicks: list[GradeClick]) -> str:
    rows = "\n".join(
        f"""
        <tr>
          <td>{click.id}</td>
          <td>{escape(click.ip_address or "-")}</td>
          <td>{escape(click.user_agent or "-")}</td>
          <td>{escape(click.clicked_at.isoformat() if click.clicked_at else "-")}</td>
        </tr>
        """
        for click in grade_clicks
    )

    if not rows:
        rows = """
        <tr>
          <td colspan="4">Записей пока нет.</td>
        </tr>
        """

    return f"""<!DOCTYPE html>
<html lang="ru">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>{escape(grade_name)} - статистика</title>
  <style>
    body {{
      margin: 0;
      font-family: "Trebuchet MS", "Segoe UI", sans-serif;
      background: linear-gradient(135deg, #eef6ff 0%, #fff8e8 100%);
      color: #1f2937;
      padding: 24px;
    }}
    .card {{
      max-width: 1100px;
      margin: 0 auto;
      background: rgba(255, 255, 255, 0.94);
      border-radius: 18px;
      padding: 28px;
      box-shadow: 0 18px 40px rgba(34, 54, 84, 0.14);
    }}
    .topbar {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 16px;
      margin-bottom: 20px;
      flex-wrap: wrap;
    }}
    h1 {{
      margin: 0;
    }}
    a {{
      color: #9f5314;
      text-decoration: none;
      font-weight: 700;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      overflow: hidden;
      border-radius: 14px;
      background: #fff;
    }}
    th, td {{
      padding: 14px 16px;
      text-align: left;
      border-bottom: 1px solid #e5e7eb;
      vertical-align: top;
    }}
    th {{
      background: #e9f1fb;
    }}
  </style>
</head>
<body>
  <div class="card">
    <div class="topbar">
      <h1>{escape(grade_name)}</h1>
      <a href="/grade-clicks/view">Назад к общей статистике</a>
    </div>
    <table>
      <thead>
        <tr>
          <th>ID</th>
          <th>IP</th>
          <th>User-Agent</th>
          <th>Время</th>
        </tr>
      </thead>
      <tbody>
        {rows}
      </tbody>
    </table>
  </div>
</body>
</html>"""


def render_not_found_page(grade_name: str) -> str:
    return f"""<!DOCTYPE html>
<html lang="ru">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Ошибка 404</title>
  <style>
    body {{
      margin: 0;
      min-height: 100vh;
      display: flex;
      align-items: center;
      justify-content: center;
      padding: 24px;
      background: radial-gradient(circle at top, #fff1df 0%, #f3f7ff 55%, #e6edf9 100%);
      font-family: "Trebuchet MS", "Segoe UI", sans-serif;
      color: #1f2937;
    }}
    .card {{
      max-width: 720px;
      width: 100%;
      background: rgba(255, 255, 255, 0.95);
      border-radius: 22px;
      padding: 36px;
      box-shadow: 0 20px 44px rgba(34, 54, 84, 0.14);
      text-align: center;
    }}
    .code {{
      font-size: 68px;
      line-height: 1;
      margin: 0 0 12px;
      color: #9f5314;
      font-weight: 800;
    }}
    h1 {{
      margin: 0 0 12px;
      font-size: 32px;
    }}
    p {{
      margin: 0 0 14px;
      color: #5b6472;
      font-size: 18px;
    }}
    a {{
      color: #9f5314;
      text-decoration: none;
      font-weight: 700;
    }}
  </style>
</head>
<body>
  <div class="card">
    <div class="code">404</div>
    <h1>Страница не найдена</h1>
    <p>Для класса {escape(grade_name)} отдельная страница сейчас недоступна.</p>
    <p>Нажатие сохранено в статистике, но запрошенный раздел отсутствует.</p>
    <a href="/">Вернуться на главную</a>
  </div>
</body>
</html>"""


@app.get("/")
def get_index():
    return FileResponse(STATIC_DIR / "index.html")


@app.get("/class-pages/{grade_name}", response_class=HTMLResponse)
def get_grade_not_found_page(grade_name: str):
    normalized_grade_name = validate_grade_name(grade_name)
    return HTMLResponse(render_not_found_page(normalized_grade_name), status_code=404)


@app.post("/visit", response_model=VisitResponse)
def create_visit(
    request: Request,
    db: Session = Depends(get_db),
):
    ip_address = get_client_ip(request)
    user_agent = request.headers.get("user-agent")

    visitor = Visitor(
        ip_address=ip_address,
        user_agent=user_agent,
    )

    db.add(visitor)
    db.commit()

    return VisitResponse(message="Visitor saved")


@app.get("/visitors", response_model=list[VisitorRecord])
def get_visitors(db: Session = Depends(get_db)):
    visitors = db.query(Visitor).order_by(desc(Visitor.visited_at), desc(Visitor.id)).all()
    return visitors


@app.get("/visitors/export")
def export_visitors_to_excel(db: Session = Depends(get_db)):
    visitors = db.query(Visitor).order_by(desc(Visitor.visited_at), desc(Visitor.id)).all()
    visits_per_ip = Counter(visitor.ip_address for visitor in visitors)
    unique_visitors: dict[str | None, Visitor] = {}

    # visitors already sorted by newest first; keep first row per ip
    for visitor in visitors:
        if visitor.ip_address not in unique_visitors:
            unique_visitors[visitor.ip_address] = visitor

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Visitors"
    worksheet.append(["id", "ip_address", "user_agent", "visited_at", "quantity"])

    for visitor in unique_visitors.values():
        visited_at = visitor.visited_at.isoformat() if visitor.visited_at else None
        quantity = visits_per_ip[visitor.ip_address]
        worksheet.append([visitor.id, visitor.ip_address, visitor.user_agent, visited_at, quantity])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"visitors_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/grade-clicks/{grade_name}", response_model=GradeClickResponse)
def create_grade_click(
    grade_name: str,
    request: Request,
    db: Session = Depends(get_db),
):
    normalized_grade_name = validate_grade_name(grade_name)
    ip_address = get_client_ip(request)
    user_agent = request.headers.get("user-agent")

    grade_click = GradeClick(
        grade_name=normalized_grade_name,
        ip_address=ip_address,
        user_agent=user_agent,
    )

    db.add(grade_click)
    db.commit()

    return GradeClickResponse(message="Grade click saved", grade_name=normalized_grade_name)


@app.get("/grade-clicks", response_model=list[GradeClickStats])
def get_grade_click_stats(db: Session = Depends(get_db)):
    rows = (
        db.query(GradeClick.grade_name, func.count(GradeClick.id).label("clicks"))
        .group_by(GradeClick.grade_name)
        .all()
    )
    clicks_map = {grade_name: clicks for grade_name, clicks in rows}
    return [GradeClickStats(grade_name=grade_name, clicks=clicks_map.get(grade_name, 0)) for grade_name in GRADE_NAMES]


@app.get("/grade-clicks/{grade_name}", response_model=list[GradeClickRecord])
def get_grade_clicks_by_grade(grade_name: str, db: Session = Depends(get_db)):
    normalized_grade_name = validate_grade_name(grade_name)
    grade_clicks = (
        db.query(GradeClick)
        .filter(GradeClick.grade_name == normalized_grade_name)
        .order_by(desc(GradeClick.clicked_at), desc(GradeClick.id))
        .all()
    )
    return grade_clicks


@app.get("/grade-clicks/export")
def export_grade_clicks_to_excel(db: Session = Depends(get_db)):
    stats = get_grade_click_stats(db)
    grade_clicks = db.query(GradeClick).order_by(desc(GradeClick.clicked_at), desc(GradeClick.id)).all()

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Grade Stats"
    summary_sheet.append(["grade_name", "clicks"])

    for item in stats:
        summary_sheet.append([item.grade_name, item.clicks])

    details_sheet = workbook.create_sheet(title="Grade Clicks")
    details_sheet.append(["id", "grade_name", "ip_address", "user_agent", "clicked_at"])

    for click in grade_clicks:
        clicked_at = click.clicked_at.isoformat() if click.clicked_at else None
        details_sheet.append([click.id, click.grade_name, click.ip_address, click.user_agent, clicked_at])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"grade_clicks_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/grade-clicks/view", response_class=HTMLResponse)
def get_grade_click_stats_view(db: Session = Depends(get_db)):
    stats = get_grade_click_stats(db)
    return HTMLResponse(render_grade_stats_page(stats))


@app.get("/grade-clicks/view/{grade_name}", response_class=HTMLResponse)
def get_grade_clicks_by_grade_view(grade_name: str, db: Session = Depends(get_db)):
    normalized_grade_name = validate_grade_name(grade_name)
    grade_clicks = get_grade_clicks_by_grade(normalized_grade_name, db)
    return HTMLResponse(render_grade_detail_page(normalized_grade_name, grade_clicks))
